from pathlib import Path
from datetime import datetime
import re

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


class TravelExcelService:

    FILE_NAME = "Reiseübersicht.xlsx"

    HEADERS = [
        "Datei",
        "Titel",
        "Dokumentart",
        "Datum",
        "Einzelkosten",
        "Währung",
        "Ort",
        "Land",
        "Reisebeginn",
        "Reiseende",
        "Kategorie",
        "Bemerkung"
    ]

    # ---------------------------------------------------------
    # Hauptfunktion
    # ---------------------------------------------------------

    def update(self, data):

        # -----------------------------------------------------
        # Reiseordner bestimmen
        # -----------------------------------------------------

        folder = Path("../Inbox") / data["folder"]

        folder.mkdir(parents=True, exist_ok=True)

        excel_path = folder / self.FILE_NAME

        # -----------------------------------------------------
        # Excel öffnen oder neu erstellen
        # -----------------------------------------------------

        if excel_path.exists():

            wb = load_workbook(excel_path)

            if "Übersicht" in wb.sheetnames:
                ws = wb["Übersicht"]
            else:
                ws = wb.create_sheet("Übersicht")

        else:

            wb = Workbook()
            ws = wb.active
            ws.title = "Übersicht"

        # -----------------------------------------------------
        # Alte Summary entfernen
        #
        # Dadurch entsteht nicht bei jedem neuen Dokument
        # ein neuer Block "Gesamtkosten" / "Kosten nach..."
        # -----------------------------------------------------

        self.remove_summary(ws)

        # -----------------------------------------------------
        # Header sicherstellen
        # -----------------------------------------------------

        self.ensure_headers(ws)

        # -----------------------------------------------------
        # Dokument hinzufügen oder vorhandenes aktualisieren
        # -----------------------------------------------------

        self.add_or_update_document(ws, data)

        # -----------------------------------------------------
        # Excel formatieren
        # -----------------------------------------------------

        self.format_document_table(ws)

        # -----------------------------------------------------
        # Reiseinformationen
        # -----------------------------------------------------

        self.create_trip_information(ws, data)

        # -----------------------------------------------------
        # Zusammenfassung
        # -----------------------------------------------------

        self.create_summary(ws)

        # -----------------------------------------------------
        # Spaltenbreiten
        # -----------------------------------------------------

        self.auto_adjust_columns(ws)

        # -----------------------------------------------------
        # Speichern
        # -----------------------------------------------------
        self.finalize_sheet(ws)
        wb.save(excel_path)

        print(f"Reiseübersicht aktualisiert: {excel_path}")

    # =========================================================
    # HEADER
    # =========================================================

    def ensure_headers(self, ws):

        # Falls bereits Daten vorhanden sind, prüfen wir,
        # ob der Header bereits korrekt existiert.

        existing_headers = [
            ws.cell(row=1, column=i).value
            for i in range(1, len(self.HEADERS) + 1)
        ]

        if existing_headers != self.HEADERS:

            # Nur dann Header schreiben, wenn die erste Zeile
            # nicht bereits korrekt aufgebaut ist.

            for column, header in enumerate(self.HEADERS, start=1):

                cell = ws.cell(
                    row=1,
                    column=column
                )

                cell.value = header

                self.style_header(cell)

    # =========================================================
    # DOKUMENT HINZUFÜGEN / AKTUALISIEREN
    # =========================================================

    def add_or_update_document(self, ws, data):

        filename = data.get("filename", "")

        specific = data.get("specific_data", {})

        # -----------------------------------------------------
        # Kosten
        # -----------------------------------------------------

        costs, currency = self.parse_cost(
            specific.get("costs", "")
        )

        # Falls der Agent die Währung separat liefert,
        # bevorzugen wir diese.

        if specific.get("currency"):
            currency = specific.get("currency")

        # -----------------------------------------------------
        # Datum
        # -----------------------------------------------------

        document_date = (
            specific.get("invoice_date")
            or specific.get("date")
            or data.get("date")
            or data.get("travel_start")
            or ""
        )

        document_date = self.parse_date(document_date)

        travel_start = self.parse_date(
            data.get("travel_start")
            or specific.get("travel_start")
            or ""
        )

        travel_end = self.parse_date(
            data.get("travel_end")
            or specific.get("travel_end")
            or ""
        )

        # -----------------------------------------------------
        # Werte
        # -----------------------------------------------------

        row_values = [
            filename,
            data.get("title", ""),
            data.get("document_type", ""),
            document_date,
            costs,
            currency,
            specific.get("city", "")
            or data.get("destination", ""),
            specific.get("country", "")
            or data.get("country", ""),
            travel_start,
            travel_end,
            data.get("travel_type", ""),
            ""
        ]

        # -----------------------------------------------------
        # Prüfen, ob Datei bereits existiert
        # -----------------------------------------------------

        existing_row = self.find_document_row(
            ws,
            filename
        )

        if existing_row:

            row = existing_row

            for column, value in enumerate(
                row_values,
                start=1
            ):
                ws.cell(
                    row=row,
                    column=column
                ).value = value

        else:

            row = ws.max_row + 1

            # Falls direkt nach dem Header geschrieben wird
            # ist das genau die gewünschte Position.

            for column, value in enumerate(
                row_values,
                start=1
            ):
                ws.cell(
                    row=row,
                    column=column
                ).value = value

        # -----------------------------------------------------
        # Markdown-Link
        # -----------------------------------------------------

        file_cell = ws.cell(
            row=row,
            column=1
        )

        if filename:

            file_cell.hyperlink = filename
            file_cell.style = "Hyperlink"

        # -----------------------------------------------------
        # Zahlenformat für Kosten
        # -----------------------------------------------------

        cost_cell = ws.cell(
            row=row,
            column=5
        )

        if isinstance(costs, (int, float)):

            cost_cell.number_format = '#,##0.00'

        # -----------------------------------------------------
        # Datumsformat
        # -----------------------------------------------------

        for column in [4, 9, 10]:

            cell = ws.cell(
                row=row,
                column=column
            )

            if isinstance(cell.value, datetime):

                cell.number_format = "DD.MM.YYYY"

    # =========================================================
    # DOKUMENT SUCHEN
    # =========================================================

    def find_document_row(self, ws, filename):

        if not filename:
            return None

        for row in range(2, ws.max_row + 1):

            value = ws.cell(
                row=row,
                column=1
            ).value

            if value == filename:
                return row

        return None

    # =========================================================
    # KOSTEN PARSEN
    # =========================================================

    def parse_cost(self, cost_string):

        if cost_string is None:
            return "", ""

        # Bereits eine Zahl
        if isinstance(cost_string, (int, float)):

            return float(cost_string), "EUR"

        cost_string = str(cost_string).strip()

        if not cost_string:
            return "", ""

        # -----------------------------------------------------
        # Währung erkennen
        # -----------------------------------------------------

        currency = ""

        if "€" in cost_string:
            currency = "EUR"

        elif "$" in cost_string:
            currency = "USD"

        elif "CHF" in cost_string.upper():
            currency = "CHF"

        elif "GBP" in cost_string.upper():
            currency = "GBP"

        # -----------------------------------------------------
        # Zahl extrahieren
        # -----------------------------------------------------

        match = re.search(
            r"[-+]?\d[\d.,]*",
            cost_string
        )

        if not match:
            return "", currency

        number = match.group(0)

        # -----------------------------------------------------
        # Deutsche Schreibweise:
        #
        # 1.234,56
        #
        # wird zu
        #
        # 1234.56
        # -----------------------------------------------------

        if "," in number and "." in number:

            number = number.replace(".", "")
            number = number.replace(",", ".")

        elif "," in number:

            number = number.replace(",", ".")

        # -----------------------------------------------------
        # In echte Zahl umwandeln
        # -----------------------------------------------------

        try:

            return float(number), currency

        except ValueError:

            return "", currency

    # =========================================================
    # DATUM PARSEN
    # =========================================================

    def parse_date(self, value):

        if not value:
            return ""

        if isinstance(value, datetime):
            return value

        value = str(value).strip()

        formats = [
            "%d.%m.%Y",
            "%Y-%m-%d",
            "%d-%m-%Y",
            "%d/%m/%Y"
        ]

        for fmt in formats:

            try:

                return datetime.strptime(
                    value,
                    fmt
                )

            except ValueError:
                continue

        return value

    # =========================================================
    # SUMMARY ENTFERNEN
    # =========================================================

    def remove_summary(self, ws):

        summary_titles = [
            "Gesamtkosten",
            "Kosten nach Dokumentart"
        ]

        rows_to_delete = []

        for row in range(
            1,
            ws.max_row + 1
        ):

            value = ws.cell(
                row=row,
                column=1
            ).value

            if value in summary_titles:

                rows_to_delete.append(row)

        # Von unten nach oben löschen,
        # damit sich die Zeilennummern nicht verschieben.

        for row in reversed(rows_to_delete):

            ws.delete_rows(row, 1)

        # -----------------------------------------------------
        # Alte leere Zeilen am Ende entfernen
        # -----------------------------------------------------

        while (
            ws.max_row > 1
            and all(
                ws.cell(
                    row=ws.max_row,
                    column=column
                ).value is None
                for column in range(
                    1,
                    len(self.HEADERS) + 1
                )
            )
        ):

            ws.delete_rows(
                ws.max_row,
                1
            )

    # =========================================================
    # REISEINFORMATIONEN
    # =========================================================

    def create_trip_information(self, ws, data):

        # Diese Informationen werden bewusst NICHT oben
        # in die Tabelle geschrieben, damit die Dokumente
        # immer direkt unter dem Header stehen.

        pass

    # =========================================================
    # SUMMARY
    # =========================================================

    def create_summary(self, ws):

        # -----------------------------------------------------
        # letzte Dokumentzeile
        # -----------------------------------------------------

        last_document_row = ws.max_row

        if last_document_row < 2:
            return

        summary_start = last_document_row + 3

        # -----------------------------------------------------
        # Gesamtkosten
        # -----------------------------------------------------

        total_title = ws.cell(
            row=summary_start,
            column=1
        )

        total_title.value = "Gesamtkosten"
        total_title.font = Font(
            bold=True,
            size=12
        )

        total_cell = ws.cell(
            row=summary_start,
            column=5
        )

        total_cell.value = (
            f"=SUM(E2:E{last_document_row})"
        )

        total_cell.number_format = '#,##0.00'

        total_cell.font = Font(
            bold=True,
            size=12
        )

        # -----------------------------------------------------
        # Kosten nach Dokumentart
        # -----------------------------------------------------

        category_title_row = summary_start + 2

        category_title = ws.cell(
            row=category_title_row,
            column=1
        )

        category_title.value = (
            "Kosten nach Dokumentart"
        )

        category_title.font = Font(
            bold=True,
            size=12
        )

        # -----------------------------------------------------
        # Dokumentarten sammeln
        # -----------------------------------------------------

        categories = set()

        for row in range(
            2,
            last_document_row + 1
        ):

            category = ws.cell(
                row=row,
                column=3
            ).value

            if category:
                categories.add(category)

        # -----------------------------------------------------
        # Kategorien ausgeben
        # -----------------------------------------------------

        row_number = category_title_row + 1

        for category in sorted(categories):

            ws.cell(
                row=row_number,
                column=1
            ).value = category

            # SUMIF auf Dokumentart
            #
            # C = Dokumentart
            # E = Kosten

            formula = (
                f'=SUMIF('
                f'C2:C{last_document_row},'
                f'"{category}",'
                f'E2:E{last_document_row}'
                f')'
            )

            cost_cell = ws.cell(
                row=row_number,
                column=5
            )

            cost_cell.value = formula
            cost_cell.number_format = '#,##0.00'

            row_number += 1

    # =========================================================
    # FORMATIERUNG
    # =========================================================

    def style_header(self, cell):

        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )

        cell.fill = PatternFill(
            "solid",
            fgColor="1F4E78"
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    # =========================================================
    # TABELLE FORMATIEREN
    # =========================================================

    def format_document_table(self, ws):

        if ws.max_row < 2:
            return

        # -----------------------------------------------------
        # Header
        # -----------------------------------------------------

        for column in range(
            1,
            len(self.HEADERS) + 1
        ):

            self.style_header(
                ws.cell(
                    row=1,
                    column=column
                )
            )

        ws.row_dimensions[1].height = 25

        # -----------------------------------------------------
        # Datenformatierung
        # -----------------------------------------------------

        thin_border = Border(
            bottom=Side(
                style="thin",
                color="D9E1F2"
            )
        )

        for row in range(
            2,
            ws.max_row + 1
        ):

            for column in range(
                1,
                len(self.HEADERS) + 1
            ):

                cell = ws.cell(
                    row=row,
                    column=column
                )

                cell.border = thin_border

                cell.alignment = Alignment(
                    vertical="center"
                )

            # Kosten
            cost_cell = ws.cell(
                row=row,
                column=5
            )

            if isinstance(
                cost_cell.value,
                (int, float)
            ):

                cost_cell.number_format = (
                    '#,##0.00'
                )

            # Datum
            for column in [4, 9, 10]:

                date_cell = ws.cell(
                    row=row,
                    column=column
                )

                if isinstance(
                    date_cell.value,
                    datetime
                ):

                    date_cell.number_format = (
                        "DD.MM.YYYY"
                    )

    # =========================================================
    # AUTOFILTER
    # =========================================================

    def add_filter(self, ws):

        if ws.max_row < 2:
            return

        ws.auto_filter.ref = (
            f"A1:{get_column_letter(len(self.HEADERS))}"
            f"{ws.max_row}"
        )

    # =========================================================
    # SPALTENBREITEN
    # =========================================================

    def auto_adjust_columns(self, ws):

        widths = {
            "A": 30,
            "B": 40,
            "C": 22,
            "D": 14,
            "E": 15,
            "F": 12,
            "G": 20,
            "H": 18,
            "I": 15,
            "J": 15,
            "K": 15,
            "L": 30
        }

        for column, width in widths.items():

            ws.column_dimensions[
                column
            ].width = width

        # -----------------------------------------------------
        # Zeilenumbruch
        # -----------------------------------------------------

        for row in ws.iter_rows():

            for cell in row:

                cell.alignment = Alignment(
                    vertical="center",
                    wrap_text=True
                )

    # =========================================================
    # FINALISIERUNG
    # =========================================================

    def finalize_sheet(self, ws):

        # Kopfzeile fixieren
        ws.freeze_panes = "A2"

        # Filter
        self.add_filter(ws)

        # Druckansicht
        ws.sheet_view.showGridLines = False

        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

        ws.sheet_properties.pageSetUpPr.fitToPage = True
