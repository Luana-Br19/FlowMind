from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

import re


class TravelExcelService:

    FILE_NAME = "Reiseübersicht.xlsx"

    HEADERS = [
        "Datei",
        "Titel",
        "Dokumentart",
        "Lieferant",
        "Datum",
        "Einzelkosten",
        "Währung",
        "Ort",
        "Land",
        "Reisebeginn",
        "Reiseende",
        "Kategorie",
        "Bemerkung"
    ]

    def update(self, data):

        folder = Path("../Inbox") / data["folder"]

        folder.mkdir(parents=True, exist_ok=True)

        excel_path = folder / self.FILE_NAME

        if excel_path.exists():
            wb = load_workbook(excel_path)
            ws = wb["Übersicht"]
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Übersicht"

            for col, header in enumerate(self.HEADERS, start=1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True)

        if self.already_exists(ws, data["filename"]):
            return

        specific = data.get("specific_data", {})

        supplier = self.extract_supplier(data)

        costs = specific.get("costs", "")

        value, currency = self.parse_cost(costs)

        ws.append([
            data.get("filename", ""),
            data.get("title", ""),
            data.get("document_type", ""),
            supplier,
            specific.get("invoice_date", ""),
            value,
            currency,
            specific.get("city", ""),
            specific.get("country", ""),
            data.get("travel_start", ""),
            data.get("travel_end", ""),
            data.get("travel_type", ""),
            ""
        ])

        for column_cells in ws.columns:
            length = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 3, 40)

        self.create_summary(ws)

        wb.save(excel_path)

    ##################################################

    def already_exists(self, ws, filename):

        for row in ws.iter_rows(min_row=2):

            if row[0].value == filename:
                return True

        return False

    ##################################################

    def parse_cost(self, cost_string):

        if not cost_string:
            return "", ""

        match = re.search(r"([\d.,]+)\s*([A-Za-z€$]*)", cost_string)

        if not match:
            return cost_string, ""

        return match.group(1), match.group(2)

    ##################################################

    def extract_supplier(self, data):

        specific = data.get("specific_data", {})

        if specific.get("accommodation"):
            return specific["accommodation"]

        transport = specific.get("transport", [])

        if transport:
            return transport[0]

        return ""

    ##################################################

    def create_summary(self, ws):

        last = ws.max_row

        summary_start = last + 3

        ws[f"A{summary_start}"] = "Gesamtkosten"
        ws[f"A{summary_start}"].font = Font(bold=True)

        ws[f"B{summary_start}"] = f"=SUM(F2:F{last})"

        summary_start += 2

        ws[f"A{summary_start}"] = "Kosten nach Dokumentart"
        ws[f"A{summary_start}"].font = Font(bold=True)

        categories = {}

        for row in ws.iter_rows(min_row=2, max_row=last):

            art = row[2].value
            value = row[5].value

            if not art:
                continue

            try:
                value = float(str(value).replace(",", "."))
            except:
                value = 0

            categories.setdefault(art, 0)
            categories[art] += value

        row_nr = summary_start + 1

        for art, total in categories.items():

            ws[f"A{row_nr}"] = art
            ws[f"B{row_nr}"] = total

            row_nr += 1
