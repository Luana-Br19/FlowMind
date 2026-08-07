from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Border,
    Side,
    Alignment
)
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from pathlib import Path
from collections import defaultdict
from datetime import datetime
import os


class TravelExcelService:

    # ---------------------------------------------------------
    # KONFIGURATION
    # ---------------------------------------------------------

    HEADERS = [
        "Datei",
        "Titel",
        "Dokumentart",
        "Datum",
        "Einzelkosten",
        "Währung",
        "Ort",
        "Land",
        "Reisebeginn",
        "Reiseende",
        "Kategorie",
        "Bemerkung"
    ]

    # Farben / Formatierung
    HEADER_FILL = PatternFill(
        fill_type="solid",
        fgColor="1F4E78"
    )

    HEADER_FONT = Font(
        color="FFFFFF",
        bold=True
    )

    TITLE_FONT = Font(
        bold=True,
        size=14
    )

    SUBTITLE_FONT = Font(
        bold=True,
        size=11
    )

    TOTAL_FONT = Font(
        bold=True,
        size=11
    )

    THIN_SIDE = Side(
        style="thin",
        color="D9E1F2"
    )

    MEDIUM_SIDE = Side(
        style="medium",
        color="1F4E78"
    )

    THIN_BORDER = Border(
        left=THIN_SIDE,
        right=THIN_SIDE,
        top=THIN_SIDE,
        bottom=THIN_SIDE
    )

    TOTAL_BORDER = Border(
        top=MEDIUM_SIDE,
        bottom=MEDIUM_SIDE
    )

    # ---------------------------------------------------------
    # INITIALISIERUNG
    # ---------------------------------------------------------

    def __init__(self, excel_path):

        self.excel_path = Path(excel_path)

        self.sheet_name = "Reiseübersicht"

    # ---------------------------------------------------------
    # HAUPTMETHODE
    # ---------------------------------------------------------

    def add_document(self, data, markdown_path=None):
        """
        Fügt ein Reisedokument zur Excel-Datei hinzu.

        Falls die Excel-Datei noch nicht existiert,
        wird sie automatisch erstellt.

        Parameters
        ----------
        data : dict
            Vom Geschäftsreise-Agenten erzeugte Daten.

        markdown_path : str, optional
            Pfad zur zugehörigen Markdown-Datei.
        """

        # -----------------------------------------------------
        # Excel öffnen oder neu erstellen
        # -----------------------------------------------------

        if self.excel_path.exists():

            wb = load_workbook(self.excel_path)

            if self.sheet_name in wb.sheetnames:
                ws = wb[self.sheet_name]
            else:
                ws = wb.create_sheet(self.sheet_name)

        else:

            wb = Workbook()

            ws = wb.active
            ws.title = self.sheet_name

        # -----------------------------------------------------
        # Header sicherstellen
        # -----------------------------------------------------

        self.ensure_header(ws)

        # -----------------------------------------------------
        # Prüfen, ob Dokument bereits existiert
        # -----------------------------------------------------

        filename = self.get_filename(data)

        existing_row = self.find_document(ws, filename)

        if existing_row:
            row_number = existing_row
        else:
            row_number = self.get_next_document_row(ws)

        # -----------------------------------------------------
        # Daten vorbereiten
        # -----------------------------------------------------

        row_data = self.build_row(
            data,
            markdown_path
        )

        # -----------------------------------------------------
        # Daten schreiben
        # -----------------------------------------------------

        for column_index, value in enumerate(
            row_data,
            start=1
        ):

            cell = ws.cell(
                row=row_number,
                column=column_index,
                value=value
            )

            cell.border = self.THIN_BORDER

            cell.alignment = Alignment(
                vertical="center"
            )

        # -----------------------------------------------------
        # Hyperlink zur Markdown-Datei
        # -----------------------------------------------------

        if markdown_path:

            file_cell = ws.cell(
                row=row_number,
                column=1
            )

            hyperlink = self.create_relative_hyperlink(
                markdown_path
            )

            if hyperlink:

                file_cell.hyperlink = hyperlink

                file_cell.style = "Hyperlink"

        # -----------------------------------------------------
        # Zahlenformat für Kosten
        # -----------------------------------------------------

        cost_cell = ws.cell(
            row=row_number,
            column=5
        )

        cost_cell.number_format = '#,##0.00'

        # -----------------------------------------------------
        # Datumsformat
        # -----------------------------------------------------

        for column in [4, 9, 10]:

            ws.cell(
                row=row_number,
                column=column
            ).number_format = "DD.MM.YYYY"

        # -----------------------------------------------------
        # Zeilenhöhe
        # -----------------------------------------------------

        ws.row_dimensions[row_number].height = 20

        # -----------------------------------------------------
        # Auswertungsbereich aktualisieren
        # -----------------------------------------------------

        self.update_summary(ws)

        # -----------------------------------------------------
        # Abschließende Formatierung
        # -----------------------------------------------------

        self.finalize_sheet(ws)

        # -----------------------------------------------------
        # Speichern
        # -----------------------------------------------------

        self.excel_path.parent.mkdir(
            parents=True,
            exist_ok=True
        )

        wb.save(self.excel_path)

        return str(self.excel_path)

    # ---------------------------------------------------------
    # HEADER
    # ---------------------------------------------------------

    def ensure_header(self, ws):

        # Wenn das Sheet leer ist
        if ws.max_row == 1 and ws.max_column == 1:

            if ws["A1"].value is None:

                for column_index, header in enumerate(
                    self.HEADERS,
                    start=1
                ):

                    cell = ws.cell(
                        row=1,
                        column=column_index,
                        value=header
                    )

                    cell.fill = self.HEADER_FILL
                    cell.font = self.HEADER_FONT
                    cell.alignment = Alignment(
                        horizontal="center",
                        vertical="center"
                    )
                    cell.border = self.THIN_BORDER

                ws.row_dimensions[1].height = 25

                return

        # Falls Header bereits existiert,
        # fehlende Header ergänzen.

        for column_index, header in enumerate(
            self.HEADERS,
            start=1
        ):

            cell = ws.cell(
                row=1,
                column=column_index
            )

            if cell.value != header:
                cell.value = header

            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )
            cell.border = self.THIN_BORDER

    # ---------------------------------------------------------
    # DATENZEILE ERSTELLEN
    # ---------------------------------------------------------

    def build_row(self, data, markdown_path=None):

        filename = self.get_filename(data)

        title = data.get(
            "title",
            ""
        )

        document_type = (
            data.get("art")
            or data.get("document_type")
            or data.get("travel_document_type")
            or data.get("travel_type")
            or ""
        )

        date = (
            data.get("date")
            or data.get("document_date")
            or data.get("datum")
            or ""
        )

        costs = self.parse_cost(
            data.get("costs", 0)
        )

        currency = (
            data.get("currency")
            or data.get("waehrung")
            or "EUR"
        )

        location = (
            data.get("location")
            or data.get("city")
            or data.get("ort")
            or ""
        )

        country = (
            data.get("country")
            or data.get("land")
            or ""
        )

        travel_start = (
            data.get("travel_start")
            or data.get("travel_begin")
            or data.get("travel_period_start")
            or ""
        )

        travel_end = (
            data.get("travel_end")
            or data.get("travel_period_end")
            or ""
        )

        category = (
            data.get("category")
            or data.get("travel_type")
            or ""
        )

        remark = (
            data.get("remark")
            or data.get("remarks")
            or data.get("summary")
            or ""
        )

        return [
            filename,
            title,
            document_type,
            self.parse_date(date),
            costs,
            currency,
            location,
            country,
            self.parse_date(travel_start),
            self.parse_date(travel_end),
            category,
            remark
        ]

    # ---------------------------------------------------------
    # DATEINAME
    # ---------------------------------------------------------

    def get_filename(self, data):

        filename = (
            data.get("filename")
            or data.get("file")
            or data.get("name")
            or ""
        )

        return filename

    # ---------------------------------------------------------
    # DOKUMENT SUCHEN
    # ---------------------------------------------------------

    def find_document(self, ws, filename):

        if not filename:
            return None

        for row in range(2, ws.max_row + 1):

            value = ws.cell(
                row=row,
                column=1
            ).value

            if value == filename:

                return row

        return None

    # ---------------------------------------------------------
    # NÄCHSTE DOKUMENTZEILE
    # ---------------------------------------------------------

    def get_next_document_row(self, ws):

        """
        Dokumente beginnen immer in Zeile 2.

        Wichtig:
        Der Auswertungsbereich wird NICHT berücksichtigt.

        Dadurch landen neue Dokumente immer direkt
        unter dem letzten Dokument.
        """

        row = 2

        while True:

            value = ws.cell(
                row=row,
                column=1
            ).value

            if value is None:

                return row

            # Falls hier bereits ein Summary-Bereich beginnt,
            # suchen wir davor nach der letzten Dokumentzeile.

            if value in (
                "Gesamtkosten",
                "Kosten nach Dokumentart"
            ):

                return row

            row += 1

    # ---------------------------------------------------------
    # KOSTEN PARSEN
    # ---------------------------------------------------------

    def parse_cost(self, value):

        if value is None:
            return 0

        if isinstance(value, (int, float)):
            return value

        value = str(value).strip()

        if not value:
            return 0

        # Währungszeichen entfernen
        value = (
            value
            .replace("€", "")
            .replace("EUR", "")
            .replace("$", "")
            .replace("USD", "")
            .strip()
        )

        # Deutsches Zahlenformat:
        # 1.234,56 -> 1234.56

        if "," in value and "." in value:

            value = value.replace(
                ".",
                ""
            )

            value = value.replace(
                ",",
                "."
            )

        elif "," in value:

            value = value.replace(
                ",",
                "."
            )

        try:

            return float(value)

        except ValueError:

            return 0

    # ---------------------------------------------------------
    # DATUM PARSEN
    # ---------------------------------------------------------

    def parse_date(self, value):

        if value is None:
            return None

        if isinstance(value, datetime):
            return value

        if not value:
            return None

        if hasattr(value, "date"):
            return value

        value = str(value).strip()

        formats = [
            "%d.%m.%Y",
            "%Y-%m-%d",
            "%d-%m-%Y",
            "%Y/%m/%d"
        ]

        for fmt in formats:

            try:

                return datetime.strptime(
                    value,
                    fmt
                )

            except ValueError:
                continue

        # Falls kein Datum erkannt wurde,
        # Originalwert behalten

        return value

    # ---------------------------------------------------------
    # HYPERLINK
    # ---------------------------------------------------------

    def create_relative_hyperlink(
        self,
        markdown_path
    ):

        if not markdown_path:
            return None

        markdown_path = Path(
            markdown_path
        )

        try:

            relative_path = os.path.relpath(
                markdown_path,
                self.excel_path.parent
            )

            return relative_path

        except ValueError:

            return str(
                markdown_path
            )

    # ---------------------------------------------------------
    # AUSWERTUNGEN
    # ---------------------------------------------------------

    def update_summary(self, ws):

        # -----------------------------------------------------
        # Alten Summary-Bereich entfernen
        # -----------------------------------------------------

        self.remove_summary(ws)

        # -----------------------------------------------------
        # Letzte Dokumentzeile bestimmen
        # -----------------------------------------------------

        last_document_row = self.get_last_document_row(
            ws
        )

        if last_document_row < 2:
            return

        # -----------------------------------------------------
        # Erste Summary-Zeile
        # -----------------------------------------------------

        summary_start = last_document_row + 2

        # -----------------------------------------------------
        # GESAMTKOSTEN
        # -----------------------------------------------------

        ws.cell(
            row=summary_start,
            column=1,
            value="Gesamtkosten"
        )

        ws.cell(
            row=summary_start,
            column=1
        ).font = self.TOTAL_FONT

        total_cell = ws.cell(
            row=summary_start,
            column=5
        )

        # WICHTIG:
        # Nur die tatsächlichen Dokumentzeilen werden
        # summiert. Keine Zwischensummen!

        total_cell.value = (
            f"=SUM(E2:E{last_document_row})"
        )

        total_cell.number_format = '#,##0.00'

        total_cell.font = self.TOTAL_FONT

        # -----------------------------------------------------
        # KOSTEN NACH DOKUMENTART
        # -----------------------------------------------------

        type_header_row = summary_start + 2

        ws.cell(
            row=type_header_row,
            column=1,
            value="Kosten nach Dokumentart"
        )

        ws.cell(
            row=type_header_row,
            column=1
        ).font = self.SUBTITLE_FONT

        # Dokumentarten aus den tatsächlichen Daten lesen

        document_types = []

        for row in range(
            2,
            last_document_row + 1
        ):

            document_type = ws.cell(
                row=row,
                column=3
            ).value

            if document_type:

                if document_type not in document_types:

                    document_types.append(
                        document_type
                    )

        # -----------------------------------------------------
        # Jede Dokumentart genau EINMAL
        # -----------------------------------------------------

        current_row = type_header_row + 1

        for document_type in document_types:

            ws.cell(
                row=current_row,
                column=1,
                value=document_type
            )

            cost_cell = ws.cell(
                row=current_row,
                column=5
            )

            # SUMIF greift ausschließlich auf den
            # tatsächlichen Dokumentbereich zu.

            cost_cell.value = (
                f'=SUMIF('
                f'C2:C{last_document_row},'
                f'A{current_row},'
                f'E2:E{last_document_row}'
                f')'
            )

            # Das obige SUMIF würde auf Spalte A vergleichen.
            # Daher korrigieren wir es direkt auf die
            # Dokumentart-Zelle in Spalte C.

            cost_cell.value = (
                f'=SUMIF('
                f'C2:C{last_document_row},'
                f'A{current_row},'
                f'E2:E{last_document_row}'
                f')'
            )

            cost_cell.number_format = '#,##0.00'

            current_row += 1

        # -----------------------------------------------------
        # Merken, bis wohin Summary geht
        # -----------------------------------------------------

        self.summary_end_row = current_row - 1

    # ---------------------------------------------------------
    # LETZTE DOKUMENTZEILE
    # ---------------------------------------------------------

    def get_last_document_row(self, ws):

        last_row = 1

        for row in range(
            2,
            ws.max_row + 1
        ):

            filename = ws.cell(
                row=row,
                column=1
            ).value

            if filename in (
                "Gesamtkosten",
                "Kosten nach Dokumentart"
            ):

                break

            if filename:

                last_row = row

        return last_row

    # ---------------------------------------------------------
    # SUMMARY ENTFERNEN
    # ---------------------------------------------------------

    def remove_summary(self, ws):

        summary_rows = []

        for row in range(
            2,
            ws.max_row + 1
        ):

            value = ws.cell(
                row=row,
                column=1
            ).value

            if value in (
                "Gesamtkosten",
                "Kosten nach Dokumentart"
            ):

                summary_rows.append(row)

        if not summary_rows:
            return

        first_summary_row = min(
            summary_rows
        )

        # Alles ab dem alten Summary löschen

        ws.delete_rows(
            first_summary_row,
            ws.max_row - first_summary_row + 1
        )

    # ---------------------------------------------------------
    # ABSCHLIESSENDE FORMATIERUNG
    # ---------------------------------------------------------

    def finalize_sheet(self, ws):

        # -----------------------------------------------------
        # Freeze Panes
        # -----------------------------------------------------

        ws.freeze_panes = "A2"

        # -----------------------------------------------------
        # Autofilter
        # -----------------------------------------------------

        ws.auto_filter.ref = (
            f"A1:L{self.get_last_document_row(ws)}"
        )

        # -----------------------------------------------------
        # Kopfzeile
        # -----------------------------------------------------

        for cell in ws[1]:

            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

            cell.border = self.THIN_BORDER

        # -----------------------------------------------------
        # Letzte relevante Zeile bestimmen
        # -----------------------------------------------------

        last_document_row = (
            self.get_last_document_row(ws)
        )

        summary_end = getattr(
            self,
            "summary_end_row",
            last_document_row
        )

        last_row = max(
            last_document_row,
            summary_end
        )

        # -----------------------------------------------------
        # Rahmen bis ganz nach unten
        # -----------------------------------------------------

        for row in range(
            1,
            last_row + 1
        ):

            for column in range(
                1,
                len(self.HEADERS) + 1
            ):

                cell = ws.cell(
                    row=row,
                    column=column
                )

                cell.border = self.THIN_BORDER

                cell.alignment = Alignment(
                    vertical="center"
                )

        # -----------------------------------------------------
        # Dokumentbereich formatieren
        # -----------------------------------------------------

        for row in range(
            2,
            last_document_row + 1
        ):

            # Kosten
            ws.cell(
                row=row,
                column=5
            ).number_format = '#,##0.00'

            # Datum
            for column in [4, 9, 10]:

                ws.cell(
                    row=row,
                    column=column
                ).number_format = "DD.MM.YYYY"

            # Kosten rechtsbündig
            ws.cell(
                row=row,
                column=5
            ).alignment = Alignment(
                horizontal="right",
                vertical="center"
            )

        # -----------------------------------------------------
        # Summary formatieren
        # -----------------------------------------------------

        for row in range(
            last_document_row + 1,
            last_row + 1
        ):

            for column in range(
                1,
                len(self.HEADERS) + 1
            ):

                cell = ws.cell(
                    row=row,
                    column=column
                )

                cell.border = self.THIN_BORDER

        # Gesamtkosten suchen

        for row in range(
            1,
            last_row + 1
        ):

            if ws.cell(
                row=row,
                column=1
            ).value == "Gesamtkosten":

                ws.cell(
                    row=row,
                    column=1
                ).font = self.TOTAL_FONT

                ws.cell(
                    row=row,
                    column=5
                ).font = self.TOTAL_FONT

                ws.cell(
                    row=row,
                    column=5
                ).number_format = '#,##0.00'

                # stärkere obere/untere Linie

                for column in range(
                    1,
                    len(self.HEADERS) + 1
                ):

                    ws.cell(
                        row=row,
                        column=column
                    ).border = self.TOTAL_BORDER

        # -----------------------------------------------------
        # Kosten nach Dokumentart hervorheben
        # -----------------------------------------------------

        for row in range(
            1,
            last_row + 1
        ):

            if ws.cell(
                row=row,
                column=1
            ).value == "Kosten nach Dokumentart":

                ws.cell(
                    row=row,
                    column=1
                ).font = self.SUBTITLE_FONT

        # -----------------------------------------------------
        # Spaltenbreiten
        # -----------------------------------------------------

        self.autofit_columns(ws)

        # -----------------------------------------------------
        # Tabellenkopf immer sichtbar
        # -----------------------------------------------------

        ws.row_dimensions[1].height = 25

    # ---------------------------------------------------------
    # SPALTENBREITEN
    # ---------------------------------------------------------

    def autofit_columns(self, ws):

        widths = {
            "A": 28,
            "B": 42,
            "C": 20,
            "D": 14,
            "E": 16,
            "F": 12,
            "G": 20,
            "H": 18,
            "I": 15,
            "J": 15,
            "K": 15,
            "L": 40
        }

        for column, width in widths.items():

            ws.column_dimensions[
                column
            ].width = width

        # Textumbruch für längere Inhalte

        for row in ws.iter_rows():

            for cell in row:

                cell.alignment = Alignment(
                    vertical="center",
                    wrap_text=True
                )

        # Kosten wieder rechtsbündig

        for row in range(
            2,
            ws.max_row + 1
        ):

            ws.cell(
                row=row,
                column=5
            ).alignment = Alignment(
                horizontal="right",
                vertical="center"
            )
